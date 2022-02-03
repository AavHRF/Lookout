# Custom Imports
from model.objects import Region

# Web Scraping Imports
import requests
from xml.etree import ElementTree

# File Operations
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import gzip
import os

# Other Imports
from sys import argv
from typing import List, Iterable
from datetime import datetime

# Lookout: Written by Aav, released under the MIT license.
# Inspired by Spyglass (https://github.com/khronion/Spyglass)
# This software is provided as-is, with no warranty or guarantee of any kind.
# You use this software at your own risk. The responsibility for complying with API terms is yours.

VERSION = "1.0.0-a"
# Lookout adheres to semantic versioning (https://semver.org/).

# A couple of globals. Yes, I know it's gross.
passworded_regions = []
founderless_regions = []


def get_region_dump(useragent: str) -> None:
    """
    Retrieves the region data dump from the NS API and writes it to a file.

    :param useragent: str: The useragent to use when making requests.
    :return: None
    """
    with requests.get(
            "https://nationstates.net/pages/regions.xml.gz",
            headers={"User-Agent": f"Lookout/{VERSION} (Run by {useragent})"}
    ) as r:
        with open("regions.xml.gz", "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                f.write(chunk)


def get_passworded(useragent: str) -> List[str]:
    """
    Retrieves a list of all passworded regions and returns it to the main function.

    :param useragent: str: A descriptive useragent for the NS API.
    :return: List[str]: A list of all passworded regions.
    """
    with requests.get(
            "https://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=password",
            headers={"User-Agent": f"Lookout/{VERSION} (Run by {useragent})"}
    ) as r:
        root = ElementTree.fromstring(r.content)
        reg = root.find("REGIONS").text
        return reg.split(",")


def get_founderless(useragent: str) -> List[str]:
    """
    Retrieves a list of all founderless regions and returns it to the main function.

    :param useragent: str: A descriptive useragent for the NS API.
    :return: List[str]: A list of all passworded regions.
    """
    with requests.get(
            "https://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=founderless",
            headers={"User-Agent": f"Lookout/{VERSION} (Run by {useragent})"}
    ) as r:
        root = ElementTree.fromstring(r.content)
        reg = root.find("REGIONS").text
        return reg.split(",")


def parse_dump() -> Iterable[Region]:
    with gzip.open("regions.xml.gz", "rb") as f:
        tree = ElementTree.parse(f)
        for elm in tree.findall("REGION"):
            retregion = Region(
                name=elm.find("NAME").text,
                wfe=elm.find("FACTBOOK").text,
                numnations=int(elm.find("NUMNATIONS").text),
                exec_delegate=False,
                exec_founder=False,
                delegate=elm.find("DELEGATE").text,
                founder=elm.find("FOUNDER").text,
                update_time=int(elm.find("LASTUPDATE").text),
                delegate_votes=int(elm.find("DELEGATEVOTES").text),
                embassies=[tag.text for tag in elm.findall("EMBASSIES")],
                _founderless=False,
                _passworded=False,
                _minorup="",
                _majorup="",
            )
            # Prevent the factbook from being parsed as an Excel formula
            if retregion.wfe[0] in ['=', '+', "-", "@"]:
                retregion.wfe = retregion.wfe[1:]
            if "X" in elm.find("DELEGATEAUTH").text:
                retregion.exec_delegate = True
            if "X" in elm.find("FOUNDERAUTH").text:
                retregion.exec_founder = True
            if elm.find("DELEGATE").text == "0":
                retregion.delegate = None
            if elm.find("FOUNDER").text == "0":
                retregion.founder = None
            yield retregion


def chelp() -> None:
    """
    A function to print the help message. Since it's invoked twice, it cuts down on code duplication.
    :return:
    """
    print(f"Lookout {VERSION}. Written by Aav. Inspired by Spyglass. Licensed under MIT.")
    print("Usage: python3 main.py [OPTIONS]")
    print("Options:")
    print("\t-h, --help\t\t\tDisplay this help message.")
    print("\t-u, --useragent\t\t\tThe useragent to use when making requests. (Required)")
    print("\t-m, --minified\t\t\tGenerates the sheet without WFE and embassies.")
    print("\t-o, --output\t\t\tThe name of the output file.")
    print("If run without any arguments, this help message is displayed.")


def main():
    # Check through the command line arguments
    if not argv:
        # You must pass at least the useragent argument
        chelp()
        # Raising SystemExit is the same as using sys.exit(). There's no reason to import it just for this.
        raise SystemExit("No arguments passed.")

    if "-h" in argv or "--help" in argv:
        chelp()
        raise SystemExit("Help message displayed.")

    if "-u" in argv or "--useragent" in argv:
        useragent = argv[argv.index("-u") + 1]
    else:
        print("You must specify a useragent.")
        raise SystemExit("No useragent specified.")

    if "-o" in argv or "--output" in argv:
        output = argv[argv.index("-o") + 1]
    else:
        output = datetime.strftime(datetime.now(), "%Y-%m-%d")

    # Download the region data dump
    get_region_dump(useragent)

    # Retrieve a list of passworded regions
    passworded_regions = get_passworded(useragent)

    # Get the founderless regions
    founderless_regions = get_founderless(useragent)

    # Parse the region dump
    regions = []
    CumulNations = 0
    for region in parse_dump():
        if region.name in passworded_regions:
            region.passworded = True
        if region.name in founderless_regions:
            region.founderless = True
        regions.append(region)
        CumulNations += region.numnations

    # Get the total number of nations that have updated by the time a region updates
    cnations = [0]
    for region in regions:
        if len(cnations) > 0:
            cnations.append(cnations[-1] + region.numnations)
        else:
            cnations.append(region.numnations)

    # Figure out how long major update is
    major_update_start = regions[0].update_time
    major_update_end = regions[-1].update_time
    major_update_length = int(major_update_end - major_update_start) # Cast to int to truncate timestamps

    # Minor update is hardcoded for now
    minor_update_length = 2640

    # Calculate update time for each region based on the total number of nations that have updated
    for n, region in zip(cnations, regions):
        t = int(n * float(minor_update_length) / CumulNations)
        region.minorup = f"{t % 60}:{(t // 60) % 60}:{t // 3600}"
        m = int(n * float(major_update_length) / CumulNations)
        region.majorup = f"{m % 60}:{(m // 60) % 60}:{m // 3600}"

    # Workbook time!
    wb = Workbook()
    ws = wb.active

    # Set up color fills
    red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow = PatternFill(start_color="FFF00", end_color="FFFF00", fill_type="solid")
    green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # These values are taken from spyglass in order to maintain compatibility with tools derived from it
    ws['L1'].value = "World"
    ws['M1'].value = "Data"
    ws['L2'].value = "Nations"
    ws['L3'].value = "Last Major"
    ws['L4'].value = "Secs/Nation"
    ws['L5'].value = "Nations/Sec"
    ws['L6'].value = "Last Minor"
    ws['L7'].value = "Secs/Nation"
    ws['L8'].value = "Nations/Sec"
    ws['L10'].value = "Lookout Version"
    ws['L11'].value = "Date Generated"
    ws['M2'].value = CumulNations
    ws['M3'].value = major_update_length
    ws['M4'].value = float(major_update_length) / CumulNations
    ws['M5'].value = 1 / (float(major_update_length) / CumulNations)
    ws['M6'].value = minor_update_length
    ws['M7'].value = minor_update_length / CumulNations
    ws['M8'].value = 1 / float(minor_update_length / CumulNations)
    ws['M10'].value = VERSION
    ws['M11'].value = datetime.strftime(datetime.now(), "%Y-%m-%d")

    # Set up the header rows (Same as Spyglass, again, for compatibility)
    ws['A1'].value = "Regions"
    ws['B1'].value = "Region Link"
    ws['C1'].value = "# Nations"
    ws['D1'].value = "Tot. Nations"
    ws['E1'].value = "Minor Upd. (est)"
    ws['F1'].value = "Major Upd. (true)"
    ws['G1'].value = "Del. Votes"
    ws['H1'].value = "Del. Endos"
    ws['I1'].value = "WFE"
    ws['J1'].value = "Embassies"

    # Save all the regions to the spreadsheet
    for i, region in enumerate(regions):
        ws.append(
            [
                region.name + "~" if region.valid_target else "*" if region.passworded else "",
                f"https://www.nationstates.net/region={region.name.replace(' ', '_')}",
                region.numnations,
                cnations[i],
                region.minorup,
                region.majorup,
                region.delegate_votes,
                region.delegate_votes - 1,
                region.wfe,
                ", ".join(region.embassies)
            ]
        )
        # Color-code the rows by target status
        ws.cell(row=i + 2, column=1).fill = red if not region.valid_target else green if not region.passworded and region.founderless else yellow

    # Save the spreadsheet
    sheet = wb["Lookout Data"]
    sheet.column_dimensions["A"].width = 45

    print("Saving spreadsheet...")

    # Save the spreadsheet
    wb.save(filename=output)
    print("Finished!")

    # Clean up
    os.remove("regions.xml.gz")


if __name__ == "__main__":
    # if __name__ main guards against the script being run as a module
    # This is important.
    main()
